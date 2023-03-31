
"""
This takes an entire transcriptome (or genomic) from NCBI and
(1) Separates them into their transcripts.
(2) Runs Scan and Fold on each transcript.
(3) Saves all intermediate files.
(4) Runs TRANSTERMHP on all ScanFold-found structures and labels them as such
(5) Runs RNABOB using a series of descriptor files for both Ribozymes and Riboswitches and labels the ScanFold-found structures

Usage:

>>>>

START IN ScanFold_working_directory

>>>>

$ python3 ScanFold-Umbrella.py filename [options]

Options:
    -s [step size]
    -w [window size]
    -r [number of randomizations, default 50]
    -t [Folding temperature, default 37]
    -type [randomization type, defualt mono]
    -p [print to screen option, default off (1)]
    --print_random [print to screen option, default off (1)]
    -f ["filter value" (integer) default -2]
    -c ["competition filter" (integer) 1 for no competition allowed or 0 for competition allowed;
          competition filtering takes much longer but yields cleaner results]
    -y ["Youden Index", or the z-score value threshold that will be used to determine which ScanFold-Structures pass]
    -e [Cell Type, 0 = prokaryotic, 1 = eukaryotic (default = 0)]
    -a [Input the annotation file in .coords format. If no annotation file is present, a default empty one will be created. If the annotation file is in genbank feature table format, use the feature_table_2_coords.py]
    -q [Run quickly - a mode to provide the user with a choice between time-efficiency and comprehensiveness of data. Default off (0)]

    The following inputs are for positive control position analyst
    -sr [starting coordinate of the RNA structure relative to the genome (5' end)]
    -er [ending coorindate of the RNA structure relative to the genome (3' end)]
    -sg [the starting point of the ScanFold relative to the genome]

Input:
    Input an NCBI FASTA transciptome file. These read as many transcripts within the same file. Often, they are denoted with the suffix
    .rna_from_genomic or a similar notation. Also, genomic files function well too, though transcriptomic data would run more quickly than a genome of the same organism.

    The annotation file is for the use of TransTermHP and should be in the .coords file. This file type is more than a little bit anachronistic. From the assembly portion of NCBI, you can
    download a genbank feature table and convert it using the feature_table_2_coords.py file and input that file into ScanFold-Umbrella.py. Note that you don't need to use an annotation file
    for this program to function, though it makes the function of TransTermHP more optimal.

Output:
    This will output a directory filled with directories named for all of the transcripts. Within those directories are .ct files for each of the designated z-scores and the excel files
    that contains all the found and labeled structures from this program.

"""

import subprocess
import ScanFold_func_Scan
import ScanFold_func_Fold
import os
import shutil
import sys
import argparse
import string
import re
import numpy as np
import RNA
import random
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor
from Bio import SeqIO
from openpyxl import Workbook
from openpyxl import load_workbook
#import meta_pk as mpk

parser = argparse.ArgumentParser()
#ScanFold-Scan Arguments
parser.add_argument('filename',  type=str, help='input filename')
parser.add_argument('-s', type=int, default = 10, help='step size')
parser.add_argument('-w', type=int, default = 120, help='window size')
parser.add_argument('-r', type=int,default = 50, help='number of randomizations')
parser.add_argument('-t', type=int, default=37,help='Folding temperature')
parser.add_argument('-type', type=str, default='mono',help='randomization type')
parser.add_argument('-p', type=str, default='off',help='print to screen option (default off:1)')
parser.add_argument('--print_random', type=str, default='off',help='print to screen option (default off)')
#ScanFold-Fold Arguments
parser.add_argument('-f', type=int, default=-2,help='filter value')
parser.add_argument('-c', type=int, default=1,help='Competition (1 for disallow competition, 0 for allow; 1 by default)')
#youden
parser.add_argument('-y', type=str, default=-2.51, help= 'Youden Index for determining viable structures')
#TRANSTERMHP
parser.add_argument('-e', type=int, default=0, help = 'type 1 if the organism is Eukaryotic, 0 (prokaryotic) is default')
parser.add_argument('-a', type=str, help = 'input the annotation file in .coords format. If no input is given, the program will run without an annotation file.')
parser.add_argument('-q', type=str, help = 'A mode of running that is more time-efficient, but less comprehensive. Deault off (0).')
#minus strand scanning
parser.add_argument('-minus', type=int, default=0, help='type 1 to scan the minus strand as well as the plus strand')

####PCPA
parser.add_argument('-sr', type=int, help="starting coordinate of the RNA structure relative to the genome (5' end)")
parser.add_argument('-er', type=int, help="ending coorindate of the RNA structure relative to the genome (3' end)")
parser.add_argument('-sg', type=int, help='the starting point of the ScanFold relative to the genome')


args = parser.parse_args()

filename = args.filename
step_size = int(args.s)
window_size = int(args.w)
randomizations = int(args.r)
temperature = int(args.t)
type = str(args.type)
print_to_screen = str(args.p)
print_random = str(args.print_random)
filter = int(args.f)
competition = int(args.c)
youden = float(args.y)
eukorprok = int(args.e)
annotation_filename = args.a
time_efficient_switch = args.q
minus_scan = args.minus

####PCPA
start_pc_rna = args.sr
end_pc_rna = args.er
start_genome = args.sg

#this describes the object 'transcript'. The entire file will be divided into these objects.
class transcript():
    def __init__(self, name, sequence, length):
        self.name = name
        self.length = ((length // window_size) * window_size)
        self.sequence = sequence[:length]

        #the following bit of code was to help decomtaminate the corrupted file name
        new_name = ''
        for i in str(self.name):
            new_name += i
        self.name = new_name

    def identify(self):
        print(self.name)
        print(self.sequence)
        print(self.length)

    def reverse_compliment(self, list_transcripts):
        name = self.name.split('.')
        name = name[0]
        print('CALLED')
        list_letters = []
        for i in self.sequence:
            list_letters.append(i)
        single_strand_compliment_not_reversed = ''
        reverse_compliment = ''
        for index, character in enumerate(self.sequence):
            if character == 'A':
                list_letters[index] == 'T'
            elif character == 'T':
                list_letters[index] = 'A'
            elif character == 'U':
                list_letters[index] = 'A'
            elif character == 'C':
                list_letters[index] = 'G'
            elif character == 'G':
                list_letters[index] = 'C'
            single_strand_compliment_not_reversed += list_letters[index-1]
        i=len(list_letters)-1
        while i >=0 :
            reverse_compliment += list_letters[i]
            i-=1
        new_name = name[:-5] +'minus'
        temp_transcript_object = transcript(new_name+'.fasta', reverse_compliment, len(reverse_compliment))    
        rev_comp_file = open(new_name+'.fasta', 'a')
        rev_comp_file.write('>'+new_name+'\n'+reverse_compliment)
        rev_comp_file.close()
        list_transcripts.append(temp_transcript_object)
        rev_comp_file = open(new_name+'.fasta', 'r')

class sf_structure:
    def __init__(self, i_most_bond, j_most_bond, nt_sequences, transcript_object):
        self.transcript_object = transcript_object
        self.i = i_most_bond
        self.j = j_most_bond
        self.five_prime_end = float(self.i) - five_prime_extension
        if self.five_prime_end < 0:
            self.five_prime_end = 1
        self.three_prime_end =  float(self.j) + three_prime_extension
        self.length = int(self.three_prime_end) - int(self.five_prime_end)
        self.sequence = nt_sequences[:self.length]
        if self.three_prime_end > self.transcript_object.length:
            self.three_prime_end = self.transcript_object.length
        #z-scores of all involved nts?
        self.label = ''
        # = 'structural overlap', 'window overlap', or 'window saddling'
        self.label_location = ''
        self.nearby_labels = ''
        self.admin_warning = ''
        self.num_pks = 0
        self.z_avg_struc = 0
        self.z_median = 0
        self.z_high = 0
        self.all_attributes = []

    def assign_all(self):
        self.all_attributes = [self.five_prime_end, self.three_prime_end, self.i, self.j, self.z_avg_struc, self.z_median, self.z_high, self.num_pks, self.label, self.label_location, self.admin_warning, self.length, self.nearby_labels, self.sequence]

    def extol(self):
        output = ''
        for num, item in enumerate(self.all_attributes):
            output += str(item)
            if (num + 1) == len(self.all_attributes):
                output += '\n'
            else:
                output += '\t'
        return output

    #the following functions/attributes of the structures add the fully labeled structures to the excel workbook in their designated sheets and rows/columns.
    def add_to_all_sheet(self, row_number):
        i = 1
        while i <= len(self.all_attributes):
            all_structures_sheet.cell(row=row_number, column=i).value = self.all_attributes[i-1]
            i += 1
    def add_to_unknown_sheet(self, row_number):
        i = 1
        while i <= len(self.all_attributes):
            all_unknown_structures_sheet.cell(row=row_number, column=i).value = self.all_attributes[i-1]
            i += 1
    def add_to_motif_sheet(self, row_number):
        i = 1
        while i <= len(self.all_attributes):
            all_known_motifs_sheet_sf.cell(row=row_number, column=i).value = self.all_attributes[i-1]
            i += 1
    def add_to_term_sheet(self, row_number):
        i = 1
        while i <= len(self.all_attributes):
            all_known_terminators_sheet_sf.cell(row=row_number, column=i).value = self.all_attributes[i-1]
            i += 1
    def add_to_unk_w_pks(self, row_number):
        i = 1
        while i <= len(self.all_attributes):
            all_unk_strucs_w_pk_sheet.cell(row=row_number, column=i).value = self.all_attributes[i-1]
            i += 1

#these functions add the found labels to a sheet in the output workbook that are not aligned with a ScanFold found structure.
def add_to_term_no_sf(row_num, line):
    values=line.split('\t')
    for index, fact in enumerate(values):
        all_known_terminators_sheet_no_sf.cell(row=row_num+2, column = index+1).value = fact
def add_to_motif_no_sf(line_list, type):
    empty_cell = False
    i = 0
    while empty_cell == False:
        if str(all_known_motifs_sheet_no_sf.cell(row=1+i,column=1).value) != 'None':
            pass
        else:
            empty_cell == True
            break
        i +=1
    row_num = i+1
    for index, fact in enumerate(line_list):
        all_known_motifs_sheet_no_sf.cell(row=row_num, column=index+1).value = fact
        row_num += 1
    all_known_motifs_sheet_no_sf.cell(row=row_num, column=3).value = type

#These variables store the calculated correction for ScanFold's bias.
three_prime_extension = 20
five_prime_extension = 20
#This variable stores the number of base pairs that scanfold will add onto the end of a found structure in search of adjacent structures that are part of the functional group
average_intermediate_size = 20
#This desginates the bp range at which Umbrella will raise an admin warning at the porximity of another found structure.
warning_range = 50

def extracting_data_from_final_partners(line, youden_index):
    line[0] = line[0].replace('nt-', '')
    line[0] = (line[0].replace(':', ''))
    #this if determines if there is a competing bp involved
    line[0] = int(line[0].replace('*', ''))
    line[0] = int(line[0])
    for num, item in enumerate(line):
        line[num] = float(item)
    if (line[0] == int(line[2]) or line[0] == int(line[1])):
        if (float(line[4]) <= float(youden_index)):
            if (line[1] != line[2]):
                line[3] = True
                if max(line[1], line[2]) > three_prime_most_j:
                    line[5] = True
                if min(line[1], line[2]) < five_prime_most_i:
                    line[5] = False
    return line

def reaching_for_bps(five_prime_bond, three_prime_bond, five_prime_line, line_data, youden_index):
    global three_prime_most_j
    global five_prime_most_i
    reach_complete = False
    five_reach = False
    while (reach_complete == False) and (line_data < len(final_partners_lines)-1):
        line_data += 1
        reach_extraction = final_partners_lines[line_data].split('\t')
        reach_extraction = extracting_data_from_final_partners(reach_extraction, youden_index)
        if reach_extraction[5] == True:
            three_prime_most_j = max(reach_extraction[1], reach_extraction[2])
            reaching_for_bps(five_prime_bond, three_prime_most_j, five_prime_line, line_data, youden_index)
            break
        else:
            if (int(max(reach_extraction[1], reach_extraction[2])) > three_prime_most_j + average_intermediate_size):
                reach_complete = True
    reach_extraction = final_partners_lines[five_prime_line].split('\t')
    reach_extraction = extracting_data_from_final_partners(reach_extraction, youden_index)
    while (reach_complete == True) and (five_prime_line > 2) and (five_reach == False):
        five_prime_line -= 1
        reach_extraction = final_partners_lines[five_prime_line].split('\t')
        reach_extraction = extracting_data_from_final_partners(reach_extraction, youden_index)
        if reach_extraction[5] == False:
            five_prime_most_i = min(reach_extraction[1], reach_extraction[2])
            reaching_for_bps(five_prime_most_i, three_prime_most_j, five_prime_line, line_data, youden_index)
            break
        elif reach_extraction[5] == True:
            three_prime_most_j = max(reach_extraction[1], reach_extraction[2])
            reaching_for_bps(five_prime_most_i, three_prime_most_j, five_prime_line, line_data, youden_index)
        else:
            if min(reach_extraction[1], reach_extraction[2]) < five_prime_most_i - average_intermediate_size:
                five_reach = True
    reaching_outputs = [five_prime_bond, three_prime_most_j, line_data, five_prime_line]
    return reaching_outputs

def find_structures(youden_index, transcript_object):
    #this list will hold all the structures that scanfold finds.
    global list_structures
    list_structures = []
    global three_prime_most_j
    global five_prime_most_i
    three_prime_most_j = 0
    five_prime_most_i = 0
    #this will be a changing variable that holds the outermost ends of a scanfold found structure
    i_and_j_bps_outermost = []
    line_data = 2
    same_nucleotide = False
    finished = False

    #this loop extracts the data from the .final_partners final_partners in a usable way
    while line_data < len(final_partners_lines[2:])-1:
        finished = False
        extract = final_partners_lines[line_data].split('\t')
        extract = extracting_data_from_final_partners(extract, youden_index)
        if extract[5] == True:
            #this asks if a structure has already been bounded
            #If it has not been bounded, it needs to be bounded
            #If the the nts are not within the bounded region, then a new bounded region needs to be made
            five_prime_most_i = min(int(extract[1]), int(extract[2]))
            three_prime_most_j = max(int(extract[1]), int(extract[2]))
            i_and_j_bps_outermost = reaching_for_bps(five_prime_most_i, three_prime_most_j, line_data, line_data, youden_index)
            line_data = i_and_j_bps_outermost[2]
            finished = True

        if finished == True:
            #this block of code finds the sequence for a found structure of scanfold
            num = 1
            sequence_of_struc = ''
            for char in transcript_object.sequence:
                if (num >= float(i_and_j_bps_outermost[0]) - five_prime_extension) and (num <= float(i_and_j_bps_outermost[1]) + three_prime_extension) and (char != '\n'):
                    sequence_of_struc += char
                if char != '\n':
                    num += 1
            if i_and_j_bps_outermost[1] > transcript_object.length:
                i_and_j_bps_outermost[1] = transcript_object.length - three_prime_extension
                sequence_of_struc = sequence_of_struc[:i_and_j_bps_outermost[1]]
            temp_structure = sf_structure(i_and_j_bps_outermost[0], i_and_j_bps_outermost[1], sequence_of_struc, transcript_object)
            list_structures.append(temp_structure)
        line_data += 1
    #This text file holds all of the structures that scanfold found and their boundaries
    sf_structures_output_file = open(transcript_object.name+'_sf_structures.txt', 'a')
    sf_structures_output_file.write(first_line_titles)
    for final_structures in list_structures:
        final_structures.assign_all()
        p = final_structures.extol()
        sf_structures_output_file.write(str(p))

def PCPA(list_structures, object, start_pc_rna, end_pc_rna):
    found_PC = False
    intermediate_dist = 0
    five_prime_addition = 0
    three_prime_addition = 0
    minus_strand = False
    start_pc_rna -= start_genome
    end_pc_rna -= start_genome
    for num, struc in enumerate(list_structures):
        if start_pc_rna > end_pc_rna:
            place_holder = start_pc_rna
            start_pc_rna = end_pc_rna
            end_pc_rna = place_holder
            minus_strand = True

        if (struc.five_prime_end in range(start_pc_rna, end_pc_rna)) or (struc.three_prime_end in range(start_pc_rna, end_pc_rna)) or (start_pc_rna in range(int(struc.five_prime_end), int(struc.three_prime_end))) or (end_pc_rna in range(int(struc.five_prime_end), int(struc.three_prime_end))):
            found_PC = True
            print(struc.i)
            #this is the raw number that should be added to the five prime end. If positive, that means a higher nt number. If negative, that means a lower nt number = shift left
            five_prime_addition = (start_pc_rna - struc.i)
            three_prime_addition = (end_pc_rna-struc.j)
            if num+1 != len(list_structures):
                if list_structures[num+1].five_prime_end in range(start_pc_rna, end_pc_rna):
                    intermediate_dist = list_structures[num+1].i - struc.j
                    three_prime_addition = (end_pc_rna-list_structures[num+1].j)
    if minus_strand == True:
        place = five_prime_addition
        five_prime_addition = three_prime_addition
        three_prime_addition = place_holder
    return [five_prime_addition, three_prime_addition, intermediate_dist]

def transterm_run(filename):
    #This ensures that the suqeunce is prokaryotic and therefor can contain rho-ind terminators
    output_filename = object.name +'_transterm_output.txt'
    output_file = open(output_filename, 'w')
    #file.close()
    
    if eukorprok == 0:
        #this makes a default and empty annotation file if there is not inputed annotation file
        if make_fake_annotation_file == True:
            annotation_filename = 'temp_fake_annotation_file.coords'
            fake_annotation_file = open(annotation_filename,'w')
            fake_annotation_file.write('fakegene1' + '\t' + '1 2 '+ object.name.split('.')[0] + '\n' + 'fakegene2'+ '\t' +str(object.length-2)+' '+str(object.length-1)+' '+object.name[0:20])
            fake_annotation_file.close()
            fake_annotation_file = open(annotation_filename,'r')
        #this runs transterm
        subprocess.run('transterm -p ../../../transterm_hp_v2.09/expterm.dat ' + object.name + ' '+ annotation_filename +  ' > '+ str(output_filename), shell=True)
        output_file.close()
        output_file = open(output_filename, 'r')
        #this next block of code converst the transterm output file into something more manageable and useable
        usable_transterm_output = open(object.name.split('.')[0]+'_transterm_output_USABLE.txt', 'w')
        usable_transterm_output.write("5' End" +'\t'+ "3' End" +'\t'+ "Name" +'\t'+ "Region Confidence" +'\t'+ "5' Tail Sequence" +'\n')
        term_lines = output_file.readlines()
        #this deletes the nonsense at the top of the output file
        term_lines = term_lines[21:]
        for index, line in enumerate(term_lines):
            temp_item = ''
            line_list = []
            #this grabs only the lines with a terminator, not a gene
            if line[0] == ' ':
                for char in line:
                    if char != ' ':
                        temp_item += char
                    elif char == ' ' and temp_item != '':
                        line_list.append(temp_item)
                        temp_item = ''
                if line_list[0] == 'TERM':
                    usable_transterm_output.write(line_list[2]  +'\t'+  line_list[4] +'\t'+ line_list[0]+' '+line_list[1] +'\t'+ line_list[7] +'\t'+ term_lines[index+1][2:16] +'\n')
        usable_transterm_output.close()
        usable_transterm_output = open(object.name.split('.')[0]+'_transterm_output_USABLE.txt', 'r')
        usable_term_lines = usable_transterm_output.readlines()[1:]
        return [usable_term_lines, annotation_filename]

def check_terminators(usable_term_lines, term_not_outported_yet, label_found):
    if eukorprok == 0:
        for number, term in enumerate(usable_term_lines):
            term_split = term.split('\t')
            if term_not_outported_yet == True:
                add_to_term_no_sf(number, term)
            #This checks to see if the terminator is within the broadened sequence of the scanfold found structure
            if (int(struc.five_prime_end) <= int(term_split[0]) <= int(struc.three_prime_end)+1) and (int(struc.five_prime_end) <= int(term_split[1])+1 <= int(struc.three_prime_end)+1):
                struc.label = term_split[2]
                struc.label_location += 'The known terminator lies totally within the sequence window predicted for the predicted structure. '
                label_found = True
            elif (int(struc.five_prime_end) <= int(term_split[0]) <= int(struc.three_prime_end)+1):
                struc.label_location += "The known terminator saddles the sequence of the predicted structure, overhanging the 3' end. "
                label_found = True
                struc.label = term_split[2]
            elif (int(struc.five_prime_end) <= int(term_split[1])+1 <= int(struc.three_prime_end)+1):
                struc.label_location += "The known terminator saddles the sequence of the predicted structure, overhanging the 5' end. "
                label_found = True
                struc.label = term_split[2]
            #This checks to see if the terminator is within the sequence of the base pairing structure of the found structure
            if (int(struc.i) <= int(term_split[0]) <= int(struc.j)+1) and (int(struc.i) <= int(term_split[1])+1 <= int(struc.j)+1):
                struc.label_location += 'The known terminator lies totally within the base-pairing structure of the predicted structure. '
                label_found = True
                struc.label = term_split[2]
            elif (int(struc.i) <= int(term_split[0]) <= int(struc.j)+1):
                struc.label_location += "The known terminator saddles the base-pairing structure of the predicted structure, overhanging the 3' end. "
                label_found = True
                struc.label = term_split[2]
            elif (int(struc.i) <= int(term_split[1])+1 <= int(struc.j)+1):
                struc.label_location += "The known terminator saddles the base-pairing structure of the predicted structure, overhanging the 5' end. "
                label_found = True
                struc.label = term_split[2]
            #This checks the edges of the structure to see if there are any nearby terminators
            if  int(term_split[1])+1 in range((int(struc.five_prime_end)-warning_range), int(struc.five_prime_end)):
                struc.nearby_labels += term_split[2] + ". "
                struc.admin_warning += '(1) There is a known terminator within '+ str(warning_range) + ' base pairs downstream the ScanFold sequence located at ' + str(struc.five_prime_end)+ ". "
            if  int(term_split[0]) in range(int(struc.three_prime_end)+warning_range, int(struc.three_prime_end)):
                struc.nearby_labels += term_split[2]+ ". "
                struc.admin_warning += '(1) There is a known terminator within '+ str(warning_range) + ' base pairs upstream the ScanFold sequence located at ' + str(struc.five_prime_end)+ ". "
            if label_found == True and time_efficient_switch == 1:
                break
        term_not_outported_yet = False
        return [term_not_outported_yet, label_found]

def check_motifs(des_filenames, BOB_not_outported_yet, label_found):
            #this runs through all the files generated for all the descriptors in the RNABOB circuit
            for name in des_filenames:
                #this formats the strings
                name = name.replace('\n', '')
                n = open(name+"_output.txt")
                #this runs through the data lines of each file
                for line in n.readlines():
                    temp_item = ''
                    line_list = []
                    for char in line:
                        if char != ' ':
                            temp_item += char
                        elif char == ' ' and temp_item != '':
                            line_list.append(temp_item)
                            temp_item = ''
                    line_list[0] = int(line_list[0])
                    line_list[1] = int(line_list[1])
                    #this outports each lines to the excel file
                    if BOB_not_outported_yet == True:
                        add_to_motif_no_sf(line_list, name)
                    #This checks to see if the motif is within the broadened sequence of the scanfold found structure
                    if (int(struc.five_prime_end) <= int(line_list[0]) <= int(struc.three_prime_end)+1) and (int(struc.five_prime_end) <= int(line_list[1])+1 <= int(struc.three_prime_end)+1):
                        struc.label = name
                        struc.label_location += 'The known motif lies totally within the sequence window predicted for the predicted structure. '
                        label_found = True
                    elif (int(struc.five_prime_end) <= int(line_list[0]) <= int(struc.three_prime_end)+1):
                        struc.label = name
                        struc.label_location += "The known motif saddles the sequence of the predicted structure, overhanging the 3' end. "
                        label_found = True
                    elif (int(struc.five_prime_end) <= int(line_list[1])+1 <= int(struc.three_prime_end)+1):
                        struc.label = name
                        struc.label_location += "The known motif saddles the sequence of the predicted structure, overhanging the 5' end. "
                        label_found = True
                    #This checks to see if the motif is within the sequence of the base pairing structure of the found structure
                    if (int(struc.i) <= int(line_list[0]) <= int(struc.j)+1) and (int(struc.i) <= int(line_list[1])+1 <= int(struc.j)+1):
                        struc.label = name
                        label_found = True
                        struc.label_location += 'The known motif lies totally within the base-pairing structure of the predicted structure. '
                    elif (int(struc.i) <= int(line_list[0]) <= int(struc.j)+1):
                        label_found = True
                        struc.label = name
                        struc.label_location += "The known motif saddles the base-pairing structure of the predicted structure, overhanging the 3' end. "
                    elif (int(struc.i) <= int(line_list[1])+1 <= int(struc.j)+1):
                        struc.label = name
                        label_found = True
                        struc.label_location += "The known motif saddles the base-pairing structure of the predicted structure, overhanging the 5' end. "
                    #This checks the edges of the structure to see if there are any nearby terminators
                    if  int(line_list[1])+1 in range(int(struc.five_prime_end)-warning_range, int(struc.five_prime_end)):
                        struc.nearby_labels += name + '. '
                        struc.admin_warning = '(1) There is a known motif within '+ str(warning_range) + ' base pairs downstream the ScanFold sequence located at ' + str(struc.five_prime_end)+ '. '
                    if  int(line_list[0]) in range(int(struc.three_prime_end)+warning_range, int(struc.three_prime_end)):
                        struc.nearby_labels += name+ '. '
                        struc.admin_warning = '(1) There is a known motif within '+ str(warning_range) + ' base pairs upstream the ScanFold sequence located at ' + str(struc.five_prime_end)+ '. '
                    if label_found == True and time_efficient_switch == 1:
                        break


                BOB_not_outported_yet = False
                return [BOB_not_outported_yet, label_found]

def final_file_renaming(object, filename, annotation_filename):
    os.rename((object.name+'_sf_structures.txt'), ('intermediate_outputs_drct/'+object.name+'_sf_structures.txt'))
    os.rename((object.name+'sf_scan.txt').replace('.fasta', ''), ('intermediate_outputs_drct/'+object.name+'sf_scan.txt').replace('.fasta', ''))
    os.rename((object.name+'sf_scan.txt.ScanFold.log.txt').replace('.fasta', ''), ('intermediate_outputs_drct/'+object.name+'sf_scan.txt.ScanFold.log.txt').replace('.fasta', ''))
    os.rename((object.name+'sf_scan.txt.ScanFold.final_partners.txt').replace('.fasta', ''), ('intermediate_outputs_drct/'+object.name+'sf_scan.txt.ScanFold.final_partners.txt').replace('.fasta', ''))
    ##os.rename('final_partners_test.bp', 'intermediate_outputs_drct/'+'final_partners_test.bp')

    os.rename(('../'+object.name+'_pipeline_output.txt'),('intermediate_outputs_drct/'+object.name+'_pipeline_output.txt'+'_pipeline_output.txt'))
  
    os.rename((object.name.split('.')[0]+'_transterm_output_USABLE.txt').replace('.fasta', ''),('intermediate_outputs_drct/'+object.name.split('.')[0]+'_transterm_output_USABLE.txt').replace('.fasta', ''))
    os.rename((object.name+'_transterm_output.txt'),('intermediate_outputs_drct/'+object.name+'_transterm_output.txt'))
    os.rename(annotation_filename, 'intermediate_outputs_drct/'+str(annotation_filename))

#this takes two preliminary structures and determines if they overlap with one another
def overlapping_structures(struc1, struc2):
    final_i = 0
    final_j = 0
    struc_with_smaller_i = struc1
    struc_with_larger_j = struc1
    overlapping = False
    struc1.five_prime_end = int(struc1.five_prime_end)
    struc1.three_prime_end = int(struc1.three_prime_end)
    struc2.five_prime_end = int(struc2.five_prime_end)
    struc2.three_prime_end = int(struc2.three_prime_end)
    if struc1.five_prime_end in range(struc2.five_prime_end, struc2.three_prime_end+1):
        overlapping = True
        struc_with_smaller_i = struc2
    if struc1.three_prime_end in range(struc2.five_prime_end, struc2.three_prime_end+1):
        overlapping = True
        struc_with_larger_j = struc2
    if struc2.five_prime_end in range(struc1.five_prime_end, struc1.three_prime_end+1):
        overlapping = True
        struc_with_smaller_i = struc1
    if struc2.three_prime_end in range(struc1.five_prime_end, struc1.three_prime_end+1):
        overlapping = True
        struc_with_larger_j = struc1
    if overlapping == True:
        final_i = min(struc1.i, struc2.i)
        final_j = max(struc1.j, struc2.j)
    return [final_i, final_j, overlapping, struc_with_smaller_i, struc_with_larger_j]

def knit_sequences(struc1, struc2, determination_of_overlap):
    list1 = []
    list2 = []
    final_sequence = ''
    for base in determination_of_overlap[3].sequence:
        final_sequence += base
    if determination_of_overlap[0] == True:
        for base in determination_of_overlap[4].sequence[(len(final_sequence)-1+determination_of_overlap[4].i-determination_of_overlap[2]):]:
            final_sequence += base
    return final_sequence

#this function finds overlapping strucutres and knits them together into one single strucutre
def structure_final_knit(list_structures):
    #loops through structures
    for number, prelim_struc in enumerate(list_structures):
        for sec_number, sec_struc in enumerate(list_structures[number:]):
            #assesses whether the two structures overlap
            determination_of_overlap = overlapping_structures(prelim_struc, sec_struc)
            if (determination_of_overlap[2] == True) and (prelim_struc != sec_struc):
                prelim_struc.i = determination_of_overlap[0]
                prelim_struc.j = determination_of_overlap[1]
                prelim_struc.five_prime_end = prelim_struc.i - five_prime_extension
                if prelim_struc.five_prime_end <= 0:
                    prelim_struc.five_prime_end =1
                prelim_struc.three_prime_end = prelim_struc.j + three_prime_extension
                if prelim_struc.three_prime_end >= prelim_struc.transcript_object.length:
                    prelim_struc.three_prime_end = prelim_struc.transcript_object.length
                #adds the two seuqences togeher
                prelim_struc.sequence = knit_sequences(prelim_struc, sec_struc, determination_of_overlap)
                prelim_struc.length = len(prelim_struc.sequence)
                prelim_struc.label = prelim_struc.label + sec_struc.label
                prelim_struc.label_location = prelim_struc.label_location+ sec_struc.label_location
                prelim_struc.nearby_labels = prelim_struc.nearby_labels + sec_struc.nearby_labels
                prelim_struc.admin_warning = prelim_struc.admin_warning + sec_struc.admin_warning
                #removes the additional, now obsolete, structure
                list_structures.remove(sec_struc)
    return list_structures

def make_struc_fasta(struc):
    temp_fasta_name = str(struc.i)+'.txt'
    temp_fasta = open(temp_fasta_name, 'a')
    temp_fasta.write('>'+str(struc.i)+'\n')
    temp_fasta.write(struc.sequence)
    return temp_fasta_name

#OBSOLETE version of a pseudoknot identifier
def nested_strucs(file_lines):
    pseudoknot_count = 0
    smallest_num = 0
    largest_num = 0
    already_nested = False
    adjacent = False
    for line in file_lines[1:]:
        indiv_elements = line.split(' ')
        indiv_elements[2].replace('\n','')
        indiv_elements[0] = int(indiv_elements[0])
        indiv_elements[2] = int(indiv_elements[2])
        if (indiv_elements[2] != 0) and (already_nested == False):
            smallest_num = indiv_elements[0]
            largest_num = indiv_elements[2]
            already_nested = True
        if (indiv_elements[2]!=0) and (already_nested == True):
            if (indiv_elements[2] == smallest_num) and (indiv_elements[0]==largest_num):
                already_nested = False
            elif (indiv_elements[2] > largest_num):
                if adjacent == False:
                    pseudoknot_count +=1
                    adjacent = True
                    continue
                if adjacent == True:
                    continue
        adjacent = False
    return pseudoknot_count

def make_accepted_pairs_list(file_lines):
    accepted_pairs = []
    for line in file_lines[2:]:
        i_and_j = []
        line_elements = line.split(' ')
        line_elements = ' '.join(line_elements).split()
        i_and_j.append(line_elements[0])
        i_and_j.append(line_elements[4])
        accepted_pairs.append(i_and_j)
    return accepted_pairs

def count_pks_in_pseudoknots(pk_nest):
    num_pks = 0
    for num_pair, pair in enumerate(pk_nest):
        if pair[0] + 1 == pk_nest[num_pair-1][0] and pair[1] -1 == pk_nest[num_pair-1][1]:
            pass
        else:
            num_pks += 1
    return num_pks

#OBSOLETE version of a pseudoknot counter
def nested_pairs(file_lines):
    pseudoknot_count = 0
    smallest_num = 0
    largest_num = 0
    running_small = 0
    running_large = 0
    adjacent = False
    known_bases = []
    for line in file_lines[1:]:
        indiv_elements = line.split(' ')
        indiv_elements[2].replace('\n','')
        indiv_elements[0] = int(indiv_elements[0])
        indiv_elements[2] = int(indiv_elements[2])
        if (indiv_elements[0] not in known_bases) and (indiv_elements[2] != 0):
            known_bases.append(indiv_elements[0])
            known_bases.append(indiv_elements[2])
            if (smallest_num == largest_num == running_small == running_small) or (indiv_elements[0]>largest_num):
                smallest_num = running_small = indiv_elements[0]
                largest_num = running_large = indiv_elements[2]
            elif (indiv_elements[0] in range(running_small, running_large+1)) and (indiv_elements[2] in range(running_small, running_large+1)):
                running_small = indiv_elements[0]
                running_large = indiv_elements[2]
            elif ((indiv_elements[0] in range(smallest_num, largest_num+1)) and (indiv_elements[2] in range(smallest_num, largest_num+1)) and (indiv_elements[0] not in range(running_small, running_large+1)) and (indiv_elements[2] not in range(running_small, running_large+1))):
                running_small = indiv_elements[0]
                running_large = indiv_elements[2]
            elif (adjacent == False):
                pseudoknot_count += 1
                adjacent = True      
                continue  
        adjacent = False
    return pseudoknot_count

def find_num_pks(filename):
    file = open(filename,'r')
    file_lines = file.readlines()
    #num_pk = nested_strucs(file_lines)
    #####num_pk = nested_pairs(file_lines)
    num_pk = count_pks_in_pseudoknots(mpk.nested_checker(make_accepted_pairs_list(file_lines)))
    return (num_pk)

def run_spot(struc, file_extension):
    #need to make a fasta for every structure
    #then, run the pks on all the structures
    #create a directory for each one of the structures, put into intermediate drct?
    #run spot_test on all the bpseq files
    #add the num of pks to an attribute of their respective classes

    os.mkdir(str(struc.i)+'.drct')
    os.chdir(str(struc.i)+'.drct')
    temp_fasta = make_struc_fasta(struc)
    ####THIS IS A HUGE PROBLEM FOR IMPLEMENTATION TO OTHER COMPUTERS. I HATE MYSELF FOR WHAT I'M ABOUT TO DO
    while os.getcwd() != '/':
        os.chdir('../')
    os.chdir('home/pfors/programs/SPOT-RNA')
    spot_rna_output_path= '/mnt/c/Users/pfors/Desktop/Research Lab/Bevilacqua Reseach Lab/ScanFilter_SNP_pipeline/Umb_d/'+file_extension+'/SPOT_RNA_files.drct/'+str(struc.i)+'.drct/'
    spot_rna_output_path_huh= '/mnt/c/Users/pfors/Desktop/Research\ Lab/Bevilacqua\ Reseach\ Lab/ScanFilter_SNP_pipeline/Umb_d/'+file_extension+'/SPOT_RNA_files.drct/'+str(struc.i)+'.drct/'
    subprocess.run('python3 SPOT-RNA.py --inputs '+spot_rna_output_path_huh+str(temp_fasta)+' --outputs '+spot_rna_output_path_huh+' --plots True', shell=True)
    os.chdir(spot_rna_output_path)
    num_pks = find_num_pks(str(struc.i)+'.bpseq')
    struc.num_pks = num_pks
    os.chdir('../')

#finds the z-score for the strucure based on the z-score of the base pairs
def z_score_for_strucs(struc, file_lines):
    list_scores_struc = []
    print(struc.five_prime_end)
    top_bound = (int(struc.three_prime_end)-1)
    if top_bound >= ((struc.five_prime_end + struc.three_prime_end)/2):
        top_bound = ((struc.five_prime_end + struc.three_prime_end)/2)
    #grabs all the bases in the structure
    for base in range(int(struc.five_prime_end), int(top_bound)):
        list_scores_bp = []
        for line in file_lines:
            line = line.split('\t')
            line = extracting_data_from_final_partners(line, youden)
            if ((int(line[1]) == base) or (int(line[2])==base)):
                list_scores_bp.append(float(line[4]))
            elif ((int(line[0])) == base):
                list_scores_bp.append(0)
        if (min(list_scores_bp)) <= youden:
            list_scores_struc.append(min(list_scores_bp))
    struc.z_avg_struc = np.mean(list_scores_struc)
    struc.z_median = np.median(list_scores_struc)
    struc.z_high = min(list_scores_struc)

#This opens the file input
file = open(filename, 'r')
lines = file.readlines()
temp_sequence = ''
list_of_transcripts = []

#this block of code determines if there is an inputted annotation file
make_fake_annotation_file = False
if annotation_filename == None and eukorprok == 0:
    make_fake_annotation_file = True
    print ('Generating a placeholder annotation file...')

working_path = '/mnt/c/Users/pfors/Desktop/Research Lab/Bevilacqua Research Lab/ScanFilter_SNP_pipeline/Umb_d/'
path_to_SF_wd = '/mnt/c/users/pfors/Desktop/Research Lab/Bevilacqua Research Lab/ScanFilter_SNP_pipeline/'

first_line_titles = "5' Boundary" + '\t' + "3' Boundary" + '\t' + "5' Bond" + '\t' +"3' Bond" + '\t' +'Avg Z'+'\t' +'Median Z'+'\t' +'High Z'+'\t'+'(uncalculated) pks'+'\t' +'Label' + '\t' +"Label Location" + '\t' +"Admin"+'\t'+'Nearby Labels'+'\t'+"Length"+'\t' +"Sequence" +'\n'

#This creates a drct based on the file input
filename1=filename.replace('.fasta', '.drct')
try:
    os.mkdir(working_path+filename1)
except:
    pass
#This enters the newly created drct
os.chdir(working_path+filename1)

print ("Reading files and parsing transcripts...")
#this begins to read the input file, which is does one line at a time whilst numbering them
for index, value in enumerate(lines):
    #this conditional ensure that the line being read is the beginning of a new transcript, which beings with a >
    if '>' in value:
        #this loops reads all of the lines after the beginning of the transcript until it reaches another transcript and compiles all the lines of bases into one sequence
        for number, sequence_line in enumerate(lines[index+1:]):
            if sequence_line[0] != '>':
                temp_sequence += sequence_line
            else:
                break
        #this loop eliminates all illeggal and problematic characters
        illegal_chars = ['|','/','\\','=','-','\n',' ','$', '>', '<', ':', '*','[',']', '(',')', ';', ',', '.', 'lcl']
        for i in value:
            if i in illegal_chars:
                value= value.replace(i, '')
        #this puts a limit on how long the transcipt name can be
        if len(value) > 20:
            value = value[0:20]
        #determines the length of a transcript, correcting for the new line characters
        temp_sequence = temp_sequence.replace('\n', '')
        length_of_transcript = len(temp_sequence)
        temp_sequence = temp_sequence[:length_of_transcript]
        print (length_of_transcript)

        #This creates an object of all the transcripts
        temp_transcript = transcript(value+'.fasta', temp_sequence, length_of_transcript)
        list_of_transcripts.append(temp_transcript)

        #this creates a new file for the transcript and inputs its name and sequence
        temp_file = open(value+'.fasta', 'a')
        temp_file.write('>'+value+'\n'+temp_sequence)
        temp_file.close()
        temp_file = open(value+'.fasta', 'r')
        temp_sequence = ''

print ('There are ' + str(len (list_of_transcripts)) +" transcript(s) in this file. Scanning...")
iteration_counter = 1
#This runs through all the transcripts in the input file

#this scans also the minus strand of the input sequence
if minus_scan == 1:
    for object in list_of_transcripts:
        if 'minus' not in object.name:
            object.reverse_compliment(list_of_transcripts)

#loops through every sequence
for object in list_of_transcripts:
    final_output = open(object.name+'_pipeline_output.txt', 'a')
    final_output.write(first_line_titles)
    #if 'minus' not in object.name:
    #    continue
    #This makes a drct of all the transcripts, puts the transcript file into its appropriate drct, and moves us into the trasncript drct
    name1 = object.name.replace('.fasta', '.drct')
    os.mkdir(working_path+filename1+'/'+name1)
    os.rename(object.name,name1+'/'+object.name)
    os.chdir(working_path+filename1+'/'+name1)
    shorter_name = str(name1).replace('.drct','')

    # the following blocks of code create an excel file with many sheets and labels all the tops rows.
    excel_book = Workbook(working_path+filename1+'/'+name1+'/'+shorter_name+'.FINAL.xlsx')
    excel_book.save(working_path+filename1+'/'+name1+'/'+shorter_name+'.FINAL.xlsx')
    excel_book = load_workbook(working_path+filename1+'/'+name1+'/'+shorter_name+'.FINAL.xlsx')
    all_structures_sheet = excel_book.active
    all_structures_sheet.title = "All Structures"
    all_unk_strucs_w_pk_sheet = excel_book.create_sheet(title = 'All Unknown Strucs with a PK')
    all_unknown_structures_sheet = excel_book.create_sheet(title = "All Unknown Structures")
    all_known_motifs_sheet_sf = excel_book.create_sheet(title = "ScanFold Predicted Motifs")
    all_known_terminators_sheet_sf = excel_book.create_sheet(title = "ScanFold Predicted Terms")
    all_known_motifs_sheet_no_sf = excel_book.create_sheet(title = "All Predicted Motifs")
    all_known_terminators_sheet_no_sf = excel_book.create_sheet(title = "All Predicted Terms")

    list_sheets = [all_structures_sheet, all_unk_strucs_w_pk_sheet, all_unknown_structures_sheet, all_known_terminators_sheet_sf,all_known_motifs_sheet_sf]
    column_labels_list = ["5' Boundary", "3' Boundary", "5' Bond", "3' Bond", "Avg Z", "Median Z", "High Z", "Number of PKs", "Label","Label Location", "Admin", "Length", "Nearby Labels", "Sequence"]

    for sheet in list_sheets:
        for place, title in enumerate(column_labels_list):
            sheet.cell(row=1, column=place+1).value = title

    column_labels_term = ["5' End", "3' End", "Name", 'Region Confidence', "5' Tail Sequence"]
    for place, title in enumerate(column_labels_term):
        all_known_terminators_sheet_no_sf.cell(row=1, column=place+1).value = title
    column_labels_BOB = ["5' End", "3' End", "Type"]
    for place, title in enumerate(column_labels_BOB):
        all_known_motifs_sheet_no_sf.cell(row=1, column=place+1).value = title

    #this runs the first part of scanfold on the transcript FASTA file.
    ScanFold_func_Scan.scan(str(object.name), step_size, window_size, randomizations, temperature, type, print_to_screen, print_random)
    print (str(iteration_counter)+ ' of ' + str(len (list_of_transcripts)) + ' has been scanned')
    #This runs scanfold fold on the scan output file which is in th etranscript file drct
    print(os.getcwd())
    ScanFold_func_Fold.fold((str(object.name)+"sf_scan.txt").replace('.fasta',''), filter, competition)
    print (str(iteration_counter)+ ' of ' + str(len (list_of_transcripts)) + ' has been folded. Scanning next transcript...')

    iteration_counter += 1

    #opens the most important file from the scanfold family
    final_partners = open((str(object.name)+"sf_scan.txt.ScanFold.final_partners.txt").replace('.fasta',''),'r')
    final_partners_lines = final_partners.readlines()

    #runs the recursive structure searching functions to determine where the bounds of structures exist in the sequence
    find_structures(youden, object)

    list_structures = structure_final_knit(list_structures)

    #this actually runs transterm
    usable_term_lines = transterm_run(object.name)
    annotation_filename = usable_term_lines[1]
    usable_term_lines = usable_term_lines[0]

    #this opens the file that stores all the names of the files
    des_file_nest = open(path_to_SF_wd+'Dependencies/RNABOB/des_file_nest.txt', 'r')
    des_filenames = des_file_nest.readlines()
    #this block runs every descriptor through RNABOB
    for name in des_filenames:
        name = name.replace('\n', '')
        temp_bob_file = open(working_path+filename1+'/'+ name1+'/'+str(object.name), 'r')
        subprocess.run("rnabob -q "+ "/mnt/c/users/pfors/Desktop/Research\ Lab/Bevilacqua\ Reseach\ Lab/ScanFilter_SNP_pipeline/Dependencies/RNABOB/"+name+".des "+ "/mnt/c/Users/pfors/Desktop/Research\ Lab/Bevilacqua\ Reseach\ Lab/work/Umb_d/"+filename1+'/'+ name1+'/'+str(object.name)+" > "+name+'_output.txt', shell=True)

    #this is where all the results are coalated
    row_number = 2
    row_number_bob = 2
    row_number_term = 2
    row_number_unk = 2
    row_number_unk_w_pk = 2
    #these booleans ensure that the data from RNABOB and transterm are only outported to the excel file once.
    term_not_outported_yet = True
    BOB_not_outported_yet = True
    #this runs through the list of all structures
    for num_struc, struc in enumerate(list_structures):
        #calculates z-score statistics for the structure
        z_score_for_strucs(struc, final_partners_lines[2:])

        label_found = False
        #this runs through all the found terminators
        term_not_outported_yet = check_terminators(usable_term_lines, term_not_outported_yet, label_found)
        label_found = term_not_outported_yet[1]
        term_not_outported_yet = term_not_outported_yet[0]

        if label_found == True and time_efficient_switch == 1:
            pass
        else:
            #checks motif outputs from RNABOB
            BOB_not_outported_yet = check_motifs(des_filenames, BOB_not_outported_yet, label_found)
            label_found = BOB_not_outported_yet[1]
            BOB_not_outported_yet = BOB_not_outported_yet[0]
        struc.assign_all()
        p = struc.extol()
        final_output.write(p)

        #this block of code checks the surrounding area for other scanfold found structures and gives admin warnings
        try:
            if int(struc.three_prime_end) + warning_range > int(list_structures[num_struc+1].five_prime_end):
                if struc.admin_warning == '':
                    struc.admin_warning += ('(1) There is a nearby predicted ScanFold structure with the structure beginning at '+str(list_structures[num_struc+1].five_prime_end))
                else:
                    struc.admin_warning += (' (2) There is a nearby predicted ScanFold structure with the structure beginning at '+str(list_structures[num_struc+1].five_prime_end))
        except:
            pass

        if struc.admin_warning != '':
            print (str(struc.admin_warning))

        try:
            os.mkdir('SPOT_RNA_files.drct')
        except:
            pass
        os.chdir('SPOT_RNA_files.drct')
        #run_spot(struc, filename1+'/'+name1)
        os.chdir('../')
        
        struc.assign_all()
        a = struc.add_to_all_sheet(row_number)
        row_number += 1
        #this adds the structure to the correct sheet
        
        if int(struc.num_pks) != 0:
            struc.add_to_unk_w_pks(row_number_unk_w_pk)
            row_number_unk_w_pk +=1
        if struc.label == '':
            struc.add_to_unknown_sheet(row_number_unk)
            row_number_unk += 1
        elif "TERM" in str(struc.label):
            struc.add_to_term_sheet(row_number_term)
            row_number_term += 1
        elif "motif" in str(struc.label_location):
            struc.add_to_motif_sheet(row_number_bob)
            row_number_bob+=1
    for struc in list_structures:
        print(struc.num_pks, 'A')

    excel_book.save(working_path+filename1+'/'+name1+'/'+shorter_name+'.FINAL.xlsx')

    os.mkdir('intermediate_outputs_drct')

    #this moves all the intermediate files and organizes them all
    for name in des_filenames:
        name = name.replace('\n', '')
        os.rename(name+'_output.txt', 'intermediate_outputs_drct/'+ name+'_output.txt')
    final_file_renaming(object, filename, annotation_filename)

    #POSIIVE CONTROL TAKE OUT FOR REAL RUNS
    os.chdir('../')


    #os.system('python3 ScanFold_PositiveControlPositionAnalyst.py ' + 'Umb_d/'+filename1+'/'+name1 + '/intermediate_outputs_drct/' + object.name+'sf_scan.txt.ScanFold.final_partners.txt'+ ' -s %s -e %s -g %s' % (start_pc_rna, end_pc_rna, start_genome))
    #pcpa_run = PCPA(list_structures, object, start_pc_rna, end_pc_rna)
    #print (pcpa_run)
os.chdir('../../')
#os.rename(filename, 'Umb_d/'+filename1+'/'+filename)

print('\a')
